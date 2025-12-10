import usb.core
import usb.util
from USB_Fonctions import detect_and_mount_usb,mount_usb_manuellement,detect_and_check_usb,usb_presente
from Fonctions_BDD import Ajouter_Consumation_SQLITE
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import io, os
from utils.logger import log_error


CWD = os.path.dirname(os.path.realpath(__file__))
MOUNT_DIR = "/mnt/usb_cle"
DB_PATH = os.path.join(CWD, "raspberry_data.db")


def log_error(message):
    import os
    CWD = os.path.dirname(os.path.realpath(__file__))
    with open(os.path.join(CWD, "errors.log") , "a", encoding="utf-8") as f:
         f.write(f"{datetime.now().isoformat()} - {message}\n")

def print_daily_summary3(p):
    try:
        print("Indice : début impression du résumé journalier")
        conn = sqlite3.connect("raspberry_data.db")
        cursor = conn.cursor()
        print("indice 2")
        date_jour = datetime.now().date()
        start_of_day = datetime.combine(date_jour, dt_time(0, 0))
        end_of_day = datetime.combine(date_jour, dt_time(23, 59, 59))
        print("indice 3")
        cursor.execute("""
            SELECT TYPE_REPAS_STR, COUNT(*) 
            FROM Consomation
            WHERE Date_Consomation BETWEEN ? AND ?
            GROUP BY TYPE_REPAS_STR
        """, (
            start_of_day.strftime("%Y-%m-%d %H:%M:%S"),
            end_of_day.strftime("%Y-%m-%d %H:%M:%S")
        ))

        results = cursor.fetchall()
        print("indice 4")
        # Initialisation
        total_tickets = 0

        # Impression de l'en-tête
        p.set(align='center', bold=True, double_height=True)
        p.text("CANTINE\n")
        p.set(align='center', bold=False, double_height=False)
        p.text(f"Resume du {date_jour.strftime('%d/%m/%Y')}\n")
        p.text("------------------------------\n")

        # Impression des lignes de repas
        p.set(align='left', bold=False)
        for type_repas_str, count in results:
            # Alignement à gauche + droite
            line = f"{type_repas_str:<30}{count:>5}\n"
            p.text(line)
            total_tickets += count

        p.text("------------------------------\n")
        p.set(bold=True)
        p.text(f"{'TOTAL':<20}{total_tickets:>5}\n")
        print("indice 5")
        conn.close()
        p.cut()
    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé journalier : {e}")


def print_month_summary(p):
    try:
        conn = sqlite3.connect("raspberry_data.db")  # Connexion SQLite
        cursor = conn.cursor()

        # Date/heure début du mois et maintenant
        start_of_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        now = datetime.now()  # Garde l'heure courante

        # Debug
        print(f"Début du mois : {start_of_month.strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Maintenant    : {now.strftime('%Y-%m-%d %H:%M:%S')}")

        # Requête SQL avec heure exacte
        cursor.execute("""
                       SELECT date (Date_Consomation) AS day, TYPE_REPAS_STR, TYPE_REPAS, COUNT (*)
                       FROM Consomation
                       WHERE Date_Consomation BETWEEN ? AND ?
                       GROUP BY day, TYPE_REPAS
                       ORDER BY day, TYPE_REPAS
                       """, (start_of_month.strftime('%Y-%m-%d %H:%M:%S'),
                             now.strftime("%Y-%m-%d %H:%M:%S")))

        results = cursor.fetchall()

        # Impression en-tête
        p.set(align='center', bold=True, double_height=True)
        p.text("RECAPITULATIF MENSUEL\n")
        p.set(align='left', bold=False, double_height=False)
        p.text(f"Periode: {start_of_month.strftime('%d/%m/%Y')} au {now.strftime('%d/%m/%Y %H:%M')}\n")
        p.text("=" * 32 + "\n")

        current_day = None
        total_month = 0

        for day, TYPE_REPAS_STR, type_repas, count in results:
            if day != current_day:
                if current_day is not None:
                    p.text("-" * 32 + "\n")
                current_day = day
                p.text(f"{datetime.strptime(day, '%Y-%m-%d').strftime('%A %d/%m')}:\n")

            p.text(f"  {TYPE_REPAS_STR:<20}{count:>10}\n")
            total_month += count

        p.text("=" * 32 + "\n")
        p.text(f"Total mois: {total_month}\n")

        #  Ecriture fichier excel
        if usb_presente():
            mount_point = detect_and_check_usb()
            if mount_point:  # Si la clé est montée et valide...
                print_daily_report_excel_usb(3, mount_point)
                p.text("\n")
                p.set(align='center')
                p.text("Rapport copié sur clé USB !\n")
            else:
                print("⚠️ Rapport non sauvegardé : clé USB absente ou non montée.")
        conn.close()
        p.cut()
    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé journalier : {e}")
        p.cut()
        conn.close()

    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé mensuel : {e}")

        def print_weekly_summary(p):
            try:
                conn = sqlite3.connect("raspberry_data.db")  # Utilisez DB_FILE pour SQLite
                cursor = conn.cursor()

                today = datetime.now().date()
                sunday = today - timedelta(days=today.weekday() + 1)
                start_of_week = datetime.combine(sunday, time(0, 0))
                end_of_week = datetime.combine(today, time(23, 59, 59))
                cursor.execute("""
                               SELECT date (Date_Consomation) as day, TYPE_REPAS_STR, TYPE_REPAS, COUNT (*)
                               FROM Consomation
                               WHERE Date_Consomation BETWEEN ? AND ?
                               GROUP BY day, TYPE_REPAS
                               ORDER BY day, TYPE_REPAS
                               """, (start_of_week.strftime("%Y-%m-%d %H:%M:%S"),
                                     end_of_week.strftime("%Y-%m-%d %H:%M:%S")))
                results = cursor.fetchall()

                p.set(align='center', bold=True, double_height=True)
                p.text("RECAPITULATIF HEBDOMADAIRE\n")
                p.set(align='left', bold=False, double_height=False)

                p.text(f"Periode: {start_of_week.strftime('%d/%m/%Y')} au {end_of_week.strftime('%d/%m/%Y')}\n")
                p.text("=" * 32 + "\n")

                current_day = None
                total_week = 0

                for day, TYPE_REPAS_STR, type_repas, count in results:
                    if day != current_day:
                        if current_day is not None:
                            p.text("-" * 32 + "\n")
                        current_day = day
                        p.text(f"{datetime.strptime(day, '%Y-%m-%d').strftime('%A %d/%m')}:\n")

                    p.text(f"  {TYPE_REPAS_STR:<20}{count:>10}\n")
                    total_week += count

                p.text("=" * 32 + "\n")
                p.text(f"Total semaine: {total_week}\n")
                #  Ecriture fichier excel
            #     if usb_presente():
            #         mount_point = detect_and_check_usb()
            #         if mount_point:  # Si la clé est montée et valide...
            #             print_daily_report_excel_usb(2, mount_point)
            #             p.text("\n")
            #             p.set(align='center')
            #             p.text("Rapport copié sur clé USB !\n")
            #         else:
            #             print("⚠️ Rapport non sauvegardé : clé USB absente ou non montée.")
                conn.close()
                p.cut()
            # except Exception as e:
            #     log_error(f"Erreur lors de l'impression du résumé journalier : {e}")



            except Exception as e:
                log_error(f"Erreur lors de l'impression du résumé hebdomadaire : {e}")

def copy_usb_report(p,type_report):
    try :
        if usb_presente():
            mount_point = detect_and_check_usb()
            if mount_point:  # Si la clé est montée et valide...
              print_daily_report_excel_usb(type_report, mount_point,0)
              p.text("\n")
              p.set(align='center')
              p.text("Rapport copié sur clé USB !\n")
        else:
            print("⚠️ Rapport non sauvegardé : clé USB absente ou non montée.")
        p.cut()

    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé journalier : {e}")


def print_ticket(user_dict, att, slot_id, slot_label, printer, jour_annee,annee, time_conso, societe):
    print('impression de ticket')

    try:
        user_id = att.user_id
        user_name = user_dict.get(user_id, "Inconnu")
        timestamp_str = att.timestamp.strftime("%Y-%m-%d %H:%M:%S")

        try:
            printer.set(align='center', bold=True, width=2, height=2)

            printer.text(f"{societe}\n\n")
            printer.text("Ticket Repas\n")

            printer.set(align='left', bold=False, width=1, height=1)

            printer.text(f"Date      : {timestamp_str}\n")
            printer.text(f"ID        : {user_id}\n")
            printer.text(f"Nom       : {user_name}\n")
            printer.text(f"Creneau   : {slot_label}\n")
            printer.text("--------------------------\n")
            printer.text(f"{datetime.now().strftime('%H:%M:%S')}\n")

            printer.cut()

            # ❌ PLUS DE printer.close() ICI

        except Exception as printer_error:
            print(">>> ERREUR IMPRESSION :", printer_error)
            log_error(f"Erreur d'impression ID {user_id} : {printer_error}")
            return


        print('Slot_ID =' + str(slot_id))
        Ajouter_Consumation_SQLITE(
            user_id, 1, slot_id, jour_annee, annee, time_conso)

    except Exception as e:
        print(">>> ERREUR print_ticket :", e)
        log_error(f"Erreur print_ticket ID {att.user_id} : {e}")


def print_weekly_summary(p):
    try:
        conn = sqlite3.connect("raspberry_data.db")  # Utilisez DB_FILE pour SQLite
        cursor = conn.cursor()

        today = datetime.now().date()
        sunday = today - timedelta(days=today.weekday() + 1)
        start_of_week = datetime.combine(sunday, dt_time(0, 0))
        end_of_week = datetime.combine(today, dt_time(23, 59, 59))
        cursor.execute("""
                       SELECT date (Date_Consomation) as day, TYPE_REPAS_STR, TYPE_REPAS, COUNT (*)
                       FROM Consomation
                       WHERE Date_Consomation BETWEEN ? AND ?
                       GROUP BY day, TYPE_REPAS
                       ORDER BY day, TYPE_REPAS
                       """, (start_of_week.strftime("%Y-%m-%d %H:%M:%S"),
                             end_of_week.strftime("%Y-%m-%d %H:%M:%S")))
        results = cursor.fetchall()

        p.set(align='center', bold=True, double_height=True)
        p.text("RECAPITULATIF HEBDOMADAIRE\n")
        p.set(align='left', bold=False, double_height=False)

        p.text(f"Periode: {start_of_week.strftime('%d/%m/%Y')} au {end_of_week.strftime('%d/%m/%Y')}\n")
        p.text("=" * 32 + "\n")

        current_day = None
        total_week = 0

        for day, TYPE_REPAS_STR, type_repas, count in results:
            if day != current_day:
                if current_day is not None:
                    p.text("-" * 32 + "\n")
                current_day = day
                p.text(f"{datetime.strptime(day, '%Y-%m-%d').strftime('%A %d/%m')}:\n")

            p.text(f"  {TYPE_REPAS_STR:<20}{count:>10}\n")
            total_week += count

        p.text("=" * 32 + "\n")
        p.text(f"Total semaine: {total_week}\n")
        #  Ecriture fichier excel
        # if usb_presente():
        #     mount_point = detect_and_check_usb()
        #     if mount_point:  # Si la clé est montée et valide...
        #         print_daily_report_excel_usb(2, mount_point)
        #         p.text("\n")
        #         p.set(align='center')
        #         p.text("Rapport copié sur clé USB !\n")
        #     else:
        #         print("⚠️ Rapport non sauvegardé : clé USB absente ou non montée.")
        conn.close()
        p.cut()
    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé journalier : {e}")



    except Exception as e:
        log_error(f"Erreur lors de l'impression du résumé hebdomadaire : {e}")


def test_printer():
    from USB_Fonctions import get_usb_printer
    """
    Teste automatiquement la première imprimante USB trouvée.
    """
    printer = get_usb_printer()
    if printer:
        try:
            printer.text("Bonjour depuis imprimante USB !\n")
            printer.cut()
            print("🖨️ Test d’impression automatique réussi !")

        except Exception as e:
            print(f"⛔ Erreur lors de l'impression : {e}")

        finally:
            # ✅ Libère le périphérique USB proprement
            usb.util.dispose_resources(printer.device)
    else:
        print("⛔ Aucune imprimante détectée.")


import io
import os
import sqlite3
from datetime import datetime, timedelta, time as dt_time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def print_daily_report_excel_usb(type_rapport, mount_point, download):
    """
    Génère un rapport Excel résumé (journalier = 1, hebdomadaire = 2, mensuel = 3)
    à partir de la base SQLite locale.

    - type_rapport : 1, 2 ou 3
    - mount_point : chemin de sauvegarde (clé USB ou chemin personnalisé)
    - download : 1 = retourne un buffer mémoire pour téléchargement
                 0 = sauvegarde sur disque dur dans mount_point

    Retourne :
    - Si download == 1 -> (buffer, filename)
    - Sinon -> (None, chemin_fichier)
    """

    print("Début génération du rapport Excel résumé...")
    now = datetime.now()

    # --- Définition des bornes temporelles selon le type de rapport ---
    if type_rapport == 1:
        date_debut = datetime.combine(now.date(), dt_time.min)
        date_fin = datetime.combine(now.date(), dt_time.max)
        titre = f"Resume des consommations du {date_debut.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_journalier_{now.strftime('%Y%m%d')}.xlsx"

    elif type_rapport == 2:
        start_of_week = now - timedelta(days=now.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        date_debut = datetime.combine(start_of_week.date(), dt_time.min)
        date_fin = datetime.combine(end_of_week.date(), dt_time.max)
        titre = f"Resume des consommations du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_hebdomadaire_{now.strftime('%Y%m%d')}.xlsx"

    elif type_rapport == 3:
        start_of_month = now.replace(day=1)
        date_debut = datetime.combine(start_of_month.date(), dt_time.min)
        date_fin = datetime.combine(now.date(), dt_time.max)
        titre = f"Resume des consommations du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_mensuel_{now.strftime('%Y%m%d')}.xlsx"
    else:
        print("❌ Type de rapport non valide")
        return None, None

    print(f"📅 Date début : {date_debut}")
    print(f"📅 Date fin   : {date_fin}")

    # --- Récupération des données ---
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT Utilisateurs.Code_Utilisateur,
                   Utilisateurs.Nom_Prenom,
                   COUNT(*) AS total
            FROM Consomation
            INNER JOIN Utilisateurs ON Utilisateurs.Code_Utilisateur = Consomation.id_utilisateur
            WHERE Date_Consomation BETWEEN ? AND ?
            GROUP BY Utilisateurs.Code_Utilisateur, Utilisateurs.Nom_Prenom
            ORDER BY Utilisateurs.Nom_Prenom ASC
        """, (
            date_debut.strftime("%Y-%m-%d %H:%M:%S"),
            date_fin.strftime("%Y-%m-%d %H:%M:%S")
        ))
        donnees = cursor.fetchall()
        conn.close()
        print(f"✅ {len(donnees)} lignes récupérées.")
    except Exception as e:
        print(f"❌ Erreur SQL : {e}")
        return None, None

    # --- Création du fichier Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Resume"

    font_titre = Font(size=14, bold=True)
    font_entete = Font(bold=True)
    border_gray = Border(left=Side(style="thin", color="CCCCCC"),
                         right=Side(style="thin", color="CCCCCC"),
                         top=Side(style="thin", color="CCCCCC"),
                         bottom=Side(style="thin", color="CCCCCC"))
    align_left = Alignment(horizontal="left", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:C1")
    ws["A1"] = titre
    ws["A1"].font = font_titre
    ws["A1"].alignment = align_left

    ws.append([]); ws.append([]); ws.append([])

    headers = ["Code Employe", "Nom et Prenom", "Quantite"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = font_entete
        cell.alignment = align_center
        cell.border = border_gray

    for code, nom_prenom, total in donnees:
        ws.append([code, nom_prenom, total])

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=3):
        for idx, cell in enumerate(row, start=1):
            cell.border = border_gray
            cell.alignment = align_center if idx != 2 else align_left

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15

    last_data_row = ws.max_row
    total_row = last_data_row + 2
    ws[f"B{total_row}"] = "Total consommations :"
    ws[f"B{total_row}"].font = font_entete
    ws[f"C{total_row}"] = f"=SUM(C5:C{last_data_row})"
    ws[f"C{total_row}"].font = font_entete
    ws[f"C{total_row}"].alignment = align_center

    # --- Téléchargement (buffer) ou sauvegarde sur disque ---
    if download == 1:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer, nom_fichier
    else:
        if not os.path.exists(mount_point):
            os.makedirs(mount_point, exist_ok=True)
        filepath = os.path.join(mount_point, nom_fichier)
        wb.save(filepath)
        return None, filepath
def print_daily_report_pdf_usb(type_rapport, mount_point, download):
    """
    Génère un rapport PDF résumé (journalier = 1, hebdomadaire = 2, mensuel = 3)
    - download = 1 -> retourne buffer + filename
    - download = 0 -> enregistre sur disque dans mount_point
    """

    print("Début génération du rapport PDF résumé...")
    now = datetime.now()

    # --- Définition des bornes de date et nom fichier ---
    if type_rapport == 1:
        date_debut = datetime.combine(now.date(), dt_time.min)
        date_fin = datetime.combine(now.date(), dt_time.max)
        titre = f"Résumé des consommations du {date_debut.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_journalier_{now.strftime('%Y%m%d')}.pdf"

    elif type_rapport == 2:
        start_of_week = now - timedelta(days=now.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        date_debut = datetime.combine(start_of_week.date(), dt_time.min)
        date_fin = datetime.combine(end_of_week.date(), dt_time.max)
        titre = f"Résumé des consommations du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_hebdomadaire_{now.strftime('%Y%m%d')}.pdf"

    elif type_rapport == 3:
        start_of_month = now.replace(day=1)
        date_debut = datetime.combine(start_of_month.date(), dt_time.min)
        date_fin = datetime.combine(now.date(), dt_time.max)
        titre = f"Résumé des consommations du {date_debut.strftime('%d/%m/%Y')} au {date_fin.strftime('%d/%m/%Y')}"
        nom_fichier = f"resume_mensuel_{now.strftime('%Y%m%d')}.pdf"

    else:
        print("❌ Type de rapport non valide")
        return None, None

    # --- Récupération des données ---
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT Utilisateurs.Code_Utilisateur,
                   Utilisateurs.Nom_Prenom,
                   COUNT(*) AS total
            FROM Consomation
            INNER JOIN Utilisateurs ON Utilisateurs.Code_Utilisateur = Consomation.id_utilisateur
            WHERE Date_Consomation BETWEEN ? AND ?
            GROUP BY Utilisateurs.Code_Utilisateur, Utilisateurs.Nom_Prenom
            ORDER BY Utilisateurs.Nom_Prenom ASC
        """, (
            date_debut.strftime("%Y-%m-%d %H:%M:%S"),
            date_fin.strftime("%Y-%m-%d %H:%M:%S")
        ))
        donnees = cursor.fetchall()
        conn.close()
        print(f"✅ {len(donnees)} lignes récupérées.")
    except Exception as e:
        print(f"❌ Erreur SQL : {e}")
        return None, None

    # --- Création du PDF ---
    buffer = io.BytesIO() if download == 1 else None
    output_path = os.path.join(mount_point, nom_fichier) if download == 0 else None

    if download == 0 and not os.path.exists(mount_point):
        os.makedirs(mount_point, exist_ok=True)

    doc = SimpleDocTemplate(buffer if download == 1 else output_path, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Titre
    elements.append(Paragraph(titre, styles['Title']))
    elements.append(Spacer(1, 20))

    # Tableau
    data = [["Code Employé", "Nom et Prénom", "Quantité"]] + list(donnees)
    table = Table(data, colWidths=[100, 250, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)

    # Total
    total_global = sum(row[2] for row in donnees)
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total consommations : <b>{total_global}</b>", styles['Heading3']))

    # Génération finale
    doc.build(elements)

    if download == 1:
        buffer.seek(0)
        return buffer, nom_fichier
    else:
        return None, output_path